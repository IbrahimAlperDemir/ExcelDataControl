# @Authors: Berkay Yasin Yavuzyiğit / İbrahim Alper Demir

import os
import traceback

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import string
import numpy
from openpyxl.reader.excel import load_workbook


class Excel:
    def __init__(self, filepath):
        self.file_name = filepath.split("\\")[-1]
        self.error_elements1 = []
        self.error_elements2 = []
        self.error_elements3 = []
        self.error_elements4 = []
        self.error_elements5 = []
        self.error_elements6 = []
        self.error_elements7 = []
        self.error_elements8 = []
        self.error_elements9 = []
        self.error_elements10 = []
        self.error_elements11 = []
        self.error_elements12 = []
        self.error_elements13 = []
        self.error_elements14 = []
        self.error_elements15 = []
        self.error_elements16 = []
        self.error_elements17 = []
        self.error_elements18 = []
        self.error_elements19 = []
        self.error_elements20 = []
        self.error_elements21 = []
        self.error_elements22 = []
        self.error_elements23 = []
        self.error_elements24 = []
        self.error_elements25 = []
        self.error_elements26 = []
        self.error_elements27 = []
        self.error_elements28 = []
        self.error_elements29 = []
        self.error_elements30 = []
        self.error_elements31 = []
        self.error_elements32 = []
        self.error_elements33 = []
        self.error_elements34 = []
        self.error_elements35 = []
        self.error_elements36 = []
        self.error_elements37 = []
        self.error_elements38 = []
        self.error_elements39 = []
        self.error_elements40 = []
        self.error_elements41 = []
        self.error_elements42 = []
        self.error_elements43 = []
        self.error_elements44 = []
        self.error_elements45 = []
        self.error_elements46 = []
        self.error_elements47 = []
        self.error_elements48 = []
        self.error_elements49 = []
        self.error_elements50 = []


class InterFace:

    def __init__(self):
        self.path = ""
        self.exceles = []
        self.directory_path = None
        self.root = tk.Tk()
        self.root.title("PALM")
        self.root.geometry("650x300")
        self.root.configure(bg="#2e2e2e")

        self.excel_path = ""  # Excel dosyasının yolu için değişken
        self.compared_excel = False

        self.title_label = tk.Label(self.root, text="PALM", font=("Helvetica", 15, "bold"),
                                    bg="#2e2e2e", fg="white")
        self.title_label.pack(pady=20)

        self.button_frame = tk.Frame(self.root, bg="#2e2e2e")
        self.button_frame.pack(pady=20)

        self.btnSelectExcel = tk.Button(self.button_frame, text="Klasörü Seç", command=self.select_excel, bg="#4a4a4a",
                                        fg="white", font=("Helvetica", 10), width=20, height=2)
        self.btnSelectExcel.grid(row=0, column=0, padx=10, pady=10)

        self.btnCompareExcel = tk.Button(self.button_frame, text="Akışları Kontrol Et", command=self.compare_excel,
                                         bg="#4a4a4a", fg="white", font=("Helvetica", 10), width=20, height=2)
        self.btnCompareExcel.grid(row=0, column=1, padx=10, pady=10)

        self.btnCreateExcel = tk.Button(self.button_frame, text="Hata Dosyasını Oluştur", command=self.create_excel,
                                        bg="#4a4a4a", fg="white", font=("Helvetica", 10), width=20, height=2)
        self.btnCreateExcel.grid(row=0, column=2, padx=10, pady=10)

        self.selected_file_label = tk.Label(self.root, text="", font=("Arial", 10), bg="#2e2e2e", fg="white")
        self.selected_file_label.pack()

        self.compared_msg = tk.Label(self.root, text="", font=("Arial", 12), bg="#2e2e2e", fg="white")
        self.compared_msg.pack(pady=20)

        # Kuralların Açıklaması
        self.kural1 = "1. Kural: Saniye olan birimler 5 ile 120 arasında olmalıdır"
        self.kural2 = "2. Kural: Dakika olan birimler 120'den küçük olmalıdır"
        self.kural3 = "3. Kural: Kontrol Sütunu Su Miktarı ise birimi 'lt' olmalıdır"
        self.kural4 = "4.Kural - Kontrol Sütunu Sıcaklık ise birimi C olmalıdır (Kontrol Sütunu Sıcaklık ise birimi °C veya dakika-saniye olmalıdır)"
        self.kural5 = "5. Kural: Sıcaklık birimi 'C' ise 31 ile 72 arasında olmalıdır"
        self.kural6 = "6. Kural: Fonksiyonsuz Mode'da - Ana Yıkama ile MFT Bloğu arasında; Sıcaklık değerleri toplamı 55 C altında olmalıdır"
        self.kural7 = "7. Kural: Kontrol Sütunu Süre ise birimi sn veya dak olmalıdır"
        self.kural8 = "8. Kural: Tahliye ile başlamalı Tahliye ile bitmeli"
        self.kural9 = "9. Kural: TABLET fonksiyonunda sıcaklık değeri Fonksiyonsuz'dan yüksek olmalı (2.SCD'de)"
        self.kural10 = "10. Kural: TABLET total süresi fonksiyonsuzdan uzun olmalı"
        self.kural11 = "11. Kural: Sirkülasyon/Tahliye adımları 15 sn den düşük olmalıdır"
        self.kural12 = "12.Kural - Akışların içinde hijyen fonksiyonu olmalı ve tablet fonksiyonu olmalıdır (makine temizleme ve ön yıkama akışları hijyen ve tablet içermez)"
        self.kural13 = "13.Kural - TABLET fonksiyonunda 2.SCD'de C olan adımlar arasında toplam sirkülasyon min 4.5dk max 8dk olmalıdır"
        #self.kural14 = "14.Kural - Adım sayısı arttıkça adıma ait süre değeri artamaz"
        self.kural15 = "15.Kural - Fonksiyon sütunu ısıtıcı ise ve kontrol sütunu süre ise ve fonksiyon sütünü dakika ise 15 dakikadan fazla olmamalı"
        self.kural16 = "16.Kural - Rej'li sütunlar hariç 2.SCD adımında Parlatıcı adımında süre 80 sn den büyük eşit olmalı"
        self.kural17 = "17.Kural - Kapı açma için maksimum süre 110sn olmalıdır"
        self.kural18 = "18.Kural - Kurutma bloğunun BosaltmaVanası/Tahliye adımı ile Su giriş vanası arasında Fan,bekleme veya Fan/klape adımlarının olması gerekmektedir"
        self.kural19 = "19.Kural - Vanalı akışlarda Parlatıcı adımında vana pozisyonu kapalı (idle) olmamalı"
        #self.kural20 = "20.Kural - Tahliye adımı sayısı su alma adımı sayısından en az bir fazla olmalıdır. (Bir fazla olma nedeni programın tahliye ile başlaması)"
        #self.kural21 = "21.Kural - Parlatıcı atılan içinde deterjan geçen adım süresi 80 sn den düşük olmamalı"
        self.kural22 = "22.Kural - SIRKULASYON / ISITICI / DETERJAN adımında 0’dan büyük değer tanımlı olmazsa deterjan kutusu açılmamaktadır, Bu nedenle bu değerin her akışta gerekli adımda 0’dan farklı olduğu bir adım olmalıdır."
        self.kural23 = "23.Kural - MSC2 akışların su alma adımlarından sonra boşaltma vanası ve bekleme olmalı ** iki su giriş arasında olsun"
        self.kural24 = "24.Kural - kapı açma hızlı 58 ve mini 30 fonlksiyonlarında 2.SCD 'de son sıcaklık değeri 50 üzerindeyse kurutmada o kadar ekstra bekleme olmalı, mini30 hariç"
        self.kural25 = "25.Kural - İki su alma adımı arasındaki toplam tahliye süresi 30 sn'den fazla 60sn'den az olmalıdır"
        self.kural26 = "26.Kural - Akış isimleri EXTENDED sayfasına ve diğer sayfalara doğru girilmeli. (excelin isim sheetlere doğru girilmeli)"
        self.kural27 = "27.Kural - Janus tahliyeler için RPM değerleri girilmelidir."
        self.kural28 = "28.Kural - Ana yıkama , 1SCD , 2.SCD bloklarında en az 1 tane 'C, dak, sn' yazan ısıtıcı adımları 0 dan farklı olacak"
        self.kural29 = "29.Kural - Asenkron Tahliyeler için RPM değeri 0 olmalıdır"
        self.kural30 = "30.Kural - Vanalılarda vanalı RPM ve pozisyon değerlerinin girilmiş olması gerekmektedir. (su giris vanası ve tahliye dışındakilere bak)"
        self.kural31 = "31.Kural - Strong ve Strong+ kombinasyonlarda SCD1 bloğunda boşaltma vanası olmalıdır."
        #self.kural32 = "32.Kural - Auto programlarda turbitity ölçüm adımları kontrol edilmeli dolu olması gerekmektedir."
        self.kural33 = "33.Kural - Rejenerasyon ve Su giriş vanası sürülüyorsa tahliye olmalıdır. (RY Bloğundandan sonraki adımda tahliye olmalı)"
        self.kural34 = "34.Kural - Rejenerasyonlu fonksiyonlarda rejenerasyon adımının dolu olması"
        self.kural35 = "35.Kural - ½ fonksiyonunun su tüketimi ve süresi fonksiyonsuzdan az olmalı"
        self.kural36 = "36.Kural - Fast süresinin fonksiyonsuzdan kısa olması ve ana yıkama sıcaklığının fonksiyonsuzdan yüksek olmalı"
        self.kural37 = "37.Kural - 60cm Vanalı ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 202 dakika olmalıdır."
        self.kural38 = "38.Kural - 60cm Vanasız ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 158 dakika olmalıdır."
        self.kural39 = "39.Kural - 45cm Vanalı ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 197 dakika olmalıdır."
        self.kural40 = "40.Kural - 45cm Vanasız ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 154 dakika olmalıdır."
        self.kural41 = "41.Kural - KV1000'li akışlarda TEPE pozisyonu olmamalı."
        self.kural42 = "42.Kural - SIRKULASYON / ISITICI , 0 harici değer , Kapalı pozisyon + S6 S7 - FAIL vermeli"
        self.kural43 = "43.Kural - MSC2 li ürünlerde, ilk su alma adımı 6.6 litreden büyük olmalı (mini 30 ve hızlı 58 ve fast45)"
        self.kural44 = "44.Kural - Silence süresi, çalışan programın 68 fazlasından az olmalı"
        self.kural45 = "45.Kural - Fonksiyonsuz + Kapı açma süresi, fonksiyonsuz süresi ile aynı olmalı"
        self.kural46 = "46.Kural - Makine temizleme kalan zaman hatalı"
        self.kural47 = "47.Kural - Ön yıkama kalan zaman hatalı"
        self.kural48 = "48.Kural - Mini30 kalan zaman hatalı"
        self.kural49 = "49.Kural - Hızlı58 kalan zaman hatalı"
        self.kural50 = "50.Kural - Auto akışlarında 0 kir 1/2 1/1 3/2 süreleri eşit olacak"




    # Sütun Numaralarını Harflere Çevir
    @staticmethod
    def column_number_to_letter(col_num):
        letters = string.ascii_uppercase
        result = ''
        while col_num >= 0:
            col_num, remainder = divmod(col_num, 26)
            result = letters[remainder] + result
            col_num -= 1
        return result

    # ----------------------------------------

    def select_excel(self):
        self.directory_path = filedialog.askdirectory()
        if self.directory_path:
            directory_name = self.directory_path.split("/")[-1]  # Get the directory name from the path
            self.selected_file_label.config(text="Seçilen Klasör: " + self.directory_path)
            self.compared_msg.config(text="")
            print("Directory selected:", self.directory_path)
        else:
            print("No directory selected.")

    def compare_excel(self):
        if self.directory_path:
            self.exceles = []
            for filename in os.listdir(self.directory_path):
                if (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(
                        ".xlsm")) and not "fail_cells" in filename:
                    filepath = os.path.join(self.directory_path, filename)
                    self.excel_path = filename
                    self.path = filepath
                    current_file = Excel(filepath)
                    self.exceles.append(current_file)
                    try:
                        excel_file = pd.ExcelFile(filepath)
                        extended_index = [index for index, value in enumerate(excel_file.sheet_names) if
                                          "EXTENDED" in value]
                        if extended_index:
                            df_selected = pd.read_excel(filepath,
                                                        sheet_name=extended_index[0])  # Read the selected Excel file

                            # Belli başlı fonksiyonların exceldeki sütunlarını bulma

                            check_fonksiyonsuz = False
                            check_tablet = False
                            check_ekstraKurutma = False
                            check_halfLoad = False
                            check_fast = False
                            check_strong = False
                            check_kapiAcma = False

                            for col in df_selected.columns:
                                if col.__contains__("Unnamed"):
                                    element = str(df_selected.iloc[0, int(col[9:])])
                                    if not check_fonksiyonsuz and "FONKSİYONSUZ (N-1)" in element:
                                        column_fonksiyonsuz = col
                                        check_fonksiyonsuz = True
                                        print(f"'Fonksiyonsuz' bulundu, sütun adı: {column_fonksiyonsuz}")

                                    elif not check_tablet and 'TABLET (T-1)' in element:
                                        check_tablet = True
                                        column_tablet = col
                                        print(f"'TABLET' bulundu, sütun adı: {column_tablet}")

                                    elif not check_ekstraKurutma and 'EKSTRA KURUTMA (T-1)' in element:
                                        check_ekstraKurutma = True
                                        column_ekstraKurutma = col
                                        print(f"'Ekstra Kurutma' bulundu, sütun adı: {column_ekstraKurutma}")

                                    elif not check_halfLoad and '1/2 (YY-1)' in element:
                                        check_halfLoad = True
                                        column_halfLoad = col
                                        print(f"'' bulundu, sütun adı: {column_halfLoad}")

                                    elif not check_fast and 'FAST (F-1)' in element:
                                        check_fast = True
                                        column_fast = col
                                        print(f"'Fast' bulundu, sütun adı: {column_fast}")

                                    elif not check_strong and 'STRONG (SW-1)' in element:
                                        check_strong = True
                                        column_strong = col
                                        print(f"'Strong' bulundu, sütun adı: {column_strong}")

                                    elif not check_kapiAcma and 'KAPI AÇMA (KA-1)' in element:
                                        check_kapiAcma = True
                                        column_kapiAcma = col
                                        print(f"'Kapı Açma' bulundu, sütun adı: {column_kapiAcma}")

                            # Block Satır Numaraları

                            onYikama_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'OY'] + 2).tolist()
                            onYikama_rowNumber = onYikama_rowNumberList[0]
                            print(onYikama_rowNumber)
                            print("OY")

                            anaYikama_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'AY'] + 2).tolist()
                            anaYikama_rowNumber = anaYikama_rowNumberList[0]
                            print(anaYikama_rowNumber)
                            print("AY")

                            mft_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'MFT'] + 2).tolist()
                            mft_rowNumber = mft_rowNumberList[0]
                            print(mft_rowNumber)
                            print("MFT")

                            sogukDurulama_rowNumberList = (
                                    df_selected.index[df_selected.iloc[:, 0] == 'SĞD'] + 2).tolist()
                            sogukDurulama_rowNumber = sogukDurulama_rowNumberList[0]
                            print(sogukDurulama_rowNumber)
                            print("SĞD")

                            extraRinse_rowNumberList = (
                                    df_selected.index[df_selected.iloc[:, 0] == 'Extra Rinse'] + 2).tolist()
                            extraRinse_rowNumber = extraRinse_rowNumberList[0]
                            print(extraRinse_rowNumber)
                            print("Extra Rinse")

                            sicakDurulama1_rowNumberList = (
                                    df_selected.index[df_selected.iloc[:, 0] == '1.SCD'] + 2).tolist()
                            sicakDurulama1_rowNumber = sicakDurulama1_rowNumberList[0]
                            print(sicakDurulama1_rowNumber)
                            print("1.SCD")

                            sicakDurulama2_rowNumberList = (
                                    df_selected.index[df_selected.iloc[:, 0] == '2.SCD'] + 2).tolist()
                            sicakDurulama2_rowNumber = sicakDurulama2_rowNumberList[0]
                            print(sicakDurulama2_rowNumber)
                            print("2.SCD")

                            dsb_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'DSB'] + 2).tolist()
                            dsb_rowNumber = dsb_rowNumberList[0]
                            print(dsb_rowNumber)
                            print("DSB")

                            rejenarasyon_rowNumberList = (
                                    df_selected.index[df_selected.iloc[:, 0] == 'RY'] + 2).tolist()
                            rejenarasyon_rowNumber = rejenarasyon_rowNumberList[0]
                            print(rejenarasyon_rowNumber)
                            print("RY")

                            kurutma_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'KU'] + 2).tolist()
                            kurutma_rowNumber = kurutma_rowNumberList[0]
                            print(kurutma_rowNumber)
                            print("KU")

                            stop_rowNumberList = (df_selected.index[df_selected.iloc[:, 0] == 'STOP'] + 2).tolist()
                            stop_rowNumber = stop_rowNumberList[0]
                            print(stop_rowNumber)
                            print("STOP")

                            firstStep = df_selected.iloc[onYikama_rowNumber - 2, 2]
                            print(firstStep)

                            lastStep = df_selected.iloc[stop_rowNumber - 3, 2]
                            print(lastStep)

                            # Block Tanımları

                            onYikama_block = df_selected.iloc[onYikama_rowNumber - 2:anaYikama_rowNumber - 2, :]

                            anaYikama_block = df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2, :]

                            ##

                            mft_block = df_selected.iloc[mft_rowNumber - 2:sogukDurulama_rowNumber - 2, :]

                            sogukDurulama_block = df_selected.iloc[sogukDurulama_rowNumber - 2:extraRinse_rowNumber - 2,
                                                  :]

                            extraRinse_block = df_selected.iloc[extraRinse_rowNumber - 2:sicakDurulama1_rowNumber - 2,
                                               :]

                            sicakDurulama1_block = df_selected.iloc[
                                                   sicakDurulama1_rowNumber - 2:sicakDurulama2_rowNumber - 2, :]

                            sicakDurulama2_block = df_selected.iloc[sicakDurulama2_rowNumber - 2:dsb_rowNumber - 2, :]

                            dsb_block = df_selected.iloc[dsb_rowNumber - 2:rejenarasyon_rowNumber - 2, :]

                            rejenarasyon_block = df_selected.iloc[rejenarasyon_rowNumber - 2:kurutma_rowNumber - 2, :]

                            kurutma_block = df_selected.iloc[kurutma_rowNumber - 2: stop_rowNumber - 2, :]

                            # Adım Sayıları

                            onYikama_adim_sayisi = anaYikama_rowNumber - onYikama_rowNumber
                            anaYikama_adim_sayisi = mft_rowNumber - anaYikama_rowNumber

                            ##

                            mft_adim_sayisi = sogukDurulama_rowNumber - mft_rowNumber
                            sogukDurulama_adim_sayisi = extraRinse_rowNumber - sogukDurulama_rowNumber
                            extraRinse_adim_sayisi = sicakDurulama1_rowNumber - extraRinse_rowNumber
                            sicakDurulama1_adim_sayisi = sicakDurulama2_rowNumber - sicakDurulama1_rowNumber
                            sicakDurulama2_adim_sayisi = dsb_rowNumber - sicakDurulama2_rowNumber
                            dsb_adim_sayisi = rejenarasyon_rowNumber - dsb_rowNumber
                            rejenarasyon_adim_sayisi = kurutma_rowNumber - rejenarasyon_rowNumber
                            kurutma_adim_sayisi = stop_rowNumber - kurutma_rowNumber

                            # Excel dosyasındaki sütun sayısı
                            num_columns = df_selected.shape[1]

                            # Degisken Sütunlarında String olan değerleri integer'a çevir
                            df_selected.iloc[:, 5:num_columns:4] = df_selected.iloc[:, 5:num_columns:4].apply(
                                pd.to_numeric,
                                errors='coerce')

                            # Birim Sütunları

                            birim_hucreleri = df_selected.iloc[:, 6:num_columns:4]
                            fonksiyonsuz_birim_hucreleri = birim_hucreleri.iloc[:, 0]

                            # Degisken Sütunları

                            degisken_hucreleri = df_selected.iloc[:, 5:num_columns:4]

                            # -------------------------------------------------------------
                            # 1. Kural - Süre birimi saniye ise 5<=değişken<=120 olmalıdır.

                            filtered_df = df_selected[df_selected.iloc[:, 3] == 'SURE']

                            # sn olan elementlerin değerlerini kontrol için yardımcı bir fonksiyon

                            def find_all_previous_elements(row):
                                for i in range(1, len(row)):
                                    if row.iloc[i] == 'sn':
                                        if not (row.iloc[i - 1] == 0 or (5 <= row.iloc[i - 1] <= 120)):
                                            row_index = row.name
                                            column_index = self.column_number_to_letter(int(row.index[i - 1][9:]))
                                            current_file.error_elements1.append(f'{row_index + 2}{column_index}')

                            # filtrelenmiş data frame'e yardımcı fonksiyonu uygula
                            filtered_df.apply(find_all_previous_elements, axis=1)

                            # ----------------------------------------------------------
                            # 2. Kural - Süre birimi dakika ise değişken<=134 olmalıdır.
                            rule_2_condition = (

                                    (df_selected.iloc[:, 3] == "SURE") &
                                    (birim_hucreleri.eq("dak").any(axis=1)) &
                                    (degisken_hucreleri.gt(134).any(axis=1))

                            )

                            if rule_2_condition.any():
                                for index, row in df_selected.iterrows():
                                    if rule_2_condition.iloc[index]:
                                        for col in range(5, num_columns, 4):
                                            if rule_2_condition.iloc[index] and (
                                                    row.iloc[col] == "dak" or row.iloc[col] > 134):
                                                col_letter = InterFace.column_number_to_letter(col)
                                                current_file.error_elements2.append(
                                                    f"{index + 2}:{col_letter}")  # Hatalı hücreleri listeye ekle

                            # ----------------------------------------------------------
                            # 3. Kural - Kontrol Sütunu Su Miktarı ise birimi lt olmalıdır
                            rule_3_condition = (

                                    (df_selected.iloc[:, 3] == "SU MIKTARI") &
                                    birim_hucreleri.ne("lt").any(axis=1)
                            )

                            if rule_3_condition.any():
                                for index, row in df_selected.iterrows():
                                    if rule_3_condition.iloc[index]:
                                        for col in range(6, num_columns, 4):
                                            if rule_3_condition.iloc[index] and (row.iloc[col] != "lt"):
                                                col_letter = InterFace.column_number_to_letter(col)
                                                current_file.error_elements3.append(f"{index + 2}:{col_letter}")

                            # ----------------------------------------------------------
                            # 4.Kural - Kontrol Sütunu Sıcaklık ise birimi C olmalıdır (Kontrol Sütunu Sıcaklık ise birimi °C veya dakika-saniye olmalıdır)
                            condition_sicaklik = df_selected.iloc[:, 3] == "SICAKLIK"
                            condition_allowed_C = birim_hucreleri.ne("C").any(axis=1)
                            condition_allowed_sn = birim_hucreleri.ne("sn").any(axis=1)
                            condition_allowed_dak = birim_hucreleri.ne("dak").any(axis=1)

                            rule_4_condition = condition_sicaklik & (
                                        condition_allowed_C | condition_allowed_dak | condition_allowed_sn)

                            filtered_rows = df_selected.index[rule_4_condition].tolist()

                            if rule_4_condition.any():
                                for index, row in df_selected.iterrows():
                                    if rule_4_condition.iloc[index]:
                                        for col in range(6, num_columns, 4):
                                            if rule_4_condition.iloc[index] and (
                                                    row.iloc[col] != "C" and row.iloc[col] != "dak" and row.iloc[
                                                col] != "sn"):
                                                col_letter = InterFace.column_number_to_letter(col)
                                                current_file.error_elements4.append(f"{index + 2}:{col_letter}")

                            # ----------------------------------------------------------
                            # 5. Kural - Sıcaklık birimi C ise 31<=değişken<=72 olmalıdır.
                            rule_5_condition = (

                                    (df_selected.iloc[:, 3] == "SICAKLIK") &
                                    (birim_hucreleri.eq("C").any(axis=1)) &

                                    (((degisken_hucreleri.gt(0) & degisken_hucreleri.lt(31)).any(axis=1)) |
                                     (degisken_hucreleri.gt(72).any(axis=1)))
                            )

                            if rule_5_condition.any():
                                for index, row in df_selected.iterrows():
                                    if rule_5_condition.iloc[index]:
                                        for col in range(5, num_columns, 4):
                                            if rule_5_condition.iloc[index] and (row.iloc[col] == "C") or (
                                                    row.iloc[col] > 75 or (0 < row.iloc[col] < 31)):
                                                col_letter = InterFace.column_number_to_letter(col)
                                                current_file.error_elements5.append(f"{index + 2}:{col_letter}")

                            # ----------------------------------------------------------
                            # 6. Kural - Fonksiyonsuz Mode'da - Ana Yıkama ile MFT Bloğu arasında; Sıcaklık değerleri toplamı 55 C altında olmalıdır (1 dk ısıtıcı çalışması = 2 derece)
                            rule_6_condition_temperature = (

                                    (df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2, 3] == "SICAKLIK") &
                                    ((df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2,
                                      2] == "SIRKULASYON / ISITICI") | (
                                             df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2,
                                             2] == "SIRKULASYON / ISITICI / DETERJAN")) &
                                    (fonksiyonsuz_birim_hucreleri.eq("C"))

                            )

                            filtered_rows = df_selected.index[rule_6_condition_temperature].tolist()
                            updated_filtered_rows = [index for index in filtered_rows]

                            total_sum_temperature = df_selected.iloc[
                                updated_filtered_rows, 5].sum()  # Sıcaklık değerlerinin toplamı (5.sütun Fonksiyonsuz değişken sütunu)

                            rule_6_condition_time = (

                                    (df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2, 3] == "SURE") &
                                    ((df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2,
                                      2] == "SIRKULASYON / ISITICI") | (
                                             df_selected.iloc[anaYikama_rowNumber - 2:mft_rowNumber - 2,
                                             2] == "SIRKULASYON / ISITICI / DETERJAN")) &
                                    (fonksiyonsuz_birim_hucreleri.eq("dak"))

                            )

                            filtered_rows = df_selected.index[rule_6_condition_time].tolist()
                            updated_filtered_rows = [index for index in filtered_rows]

                            total_sum_time = df_selected.iloc[
                                updated_filtered_rows, 5].sum()  # Süre değerlerinin toplamı
                            sumTime_to_temperature = 2 * total_sum_time  # 1 dk ısıtıcının çalışması 2 C sıcaklığa denk gelir

                            total_temparature = total_sum_temperature + sumTime_to_temperature

                            if total_temparature <= 55:
                                current_file.error_elements6.append("HATA")

                            # ----------------------------------------------------------
                            # 7. Kural - Kontrol Sütunu Süre ise birimi sn veya dak olmalıdır
                            rule_7_condition = (
                                    (df_selected.iloc[:, 3] == "SURE") &
                                    (birim_hucreleri.eq("C").any(axis=1) |
                                     birim_hucreleri.eq("lt").any(axis=1))
                            )

                            filtered_rows = df_selected.index[rule_7_condition].tolist()
                            updated_filtered_rows = [index for index in filtered_rows]
                            print(df_selected.iloc[updated_filtered_rows, 6])

                            if rule_7_condition.any():
                                for index, row in df_selected.iterrows():
                                    if rule_7_condition.iloc[index]:
                                        for col in range(6, num_columns, 4):
                                            if rule_7_condition.iloc[index] and (
                                                    (row.iloc[col] != "sn") and (row.iloc[col] != "dak")):
                                                col_letter = InterFace.column_number_to_letter(col)
                                                current_file.error_elements7.append(f"{index + 2}:{col_letter}")

                            # ----------------------------------------------------------
                            # 8. Kural - Tahliye ile başlamalı Tahliye ile bitmeli

                            if firstStep == "TAHLIYE" and lastStep == "TAHLIYE":
                                pass
                            else:
                                current_file.error_elements8.append("HATA")

                            # ----------------------------------------------------------
                            # 9. Kural - TABLET fonksiyonunda sıcaklık değeri Fonksiyonsuz'dan yüksek olmalı (2.SCD'de)

                            # FONKSİYONSUZ ve TABLET tanımlı mı kontrolü
                            if check_fonksiyonsuz and check_tablet:

                                # FONKSİYONSUZ ve TABLET sütunlarının indexlerini bulma
                                int_column_fonksiyonsuz = int(column_fonksiyonsuz[9:])
                                int_column_tablet = int(column_tablet[9:])

                                # SCD2 Bloğunun FONKSİYONSUZ ve TABLET kısımlarını çıkarma
                                sicakDurulama2_fonksiyonsuz_block = sicakDurulama2_block.iloc[:, [2, 3,
                                                                                                  int_column_fonksiyonsuz,
                                                                                                  int_column_fonksiyonsuz + 1,
                                                                                                  int_column_fonksiyonsuz + 2]]
                                sicakDurulama2_tablet_block = sicakDurulama2_block.iloc[:, [2, 3,
                                                                                            int_column_tablet,
                                                                                            int_column_tablet + 1,
                                                                                            int_column_tablet + 2]]

                                # kontrol edilecek değerler için uygun koşullar
                                condition_ısıtıcı = sicakDurulama2_tablet_block.iloc[:, 0].str.contains("ISITICI")
                                condition_sıcaklık = sicakDurulama2_tablet_block.iloc[:, 1] == "SICAKLIK"
                                condition_c = sicakDurulama2_tablet_block.iloc[:, 4] == "C"
                                condition_cc = sicakDurulama2_fonksiyonsuz_block.iloc[:, 4] == "C"
                                condition_tablet = condition_ısıtıcı & condition_sıcaklık & condition_c
                                condition_fonksiyonsuz = condition_ısıtıcı & condition_sıcaklık & condition_cc

                                # istenenlere göre filtrelenmiş blocklar
                                filtered_sicakDurulama2_fonksiyonsuz_block = sicakDurulama2_fonksiyonsuz_block[
                                    condition_fonksiyonsuz]
                                filtered_sicakDurulama2_tablet_block = sicakDurulama2_tablet_block[condition_tablet]

                                # tablet sıcaklık değerleri daha büyük mü kontrolü
                                is_greater = filtered_sicakDurulama2_tablet_block.iloc[:,
                                             3].sum() > filtered_sicakDurulama2_fonksiyonsuz_block.iloc[:, 3].sum()

                                if is_greater:
                                    pass
                                else:
                                    current_file.error_elements9.append("HATA")

                            # ----------------------------------------------------------
                            # 10. Kural - TABLET total süresi fonksiyonsuzdan uzun olmalı

                            # FONKSİYONSUZ ve TABLET tanımlı mı kontrolü
                            if check_fonksiyonsuz and check_tablet:
                                # fonksiyonsuz ve tablet kalan zaman için excel'in en üstündeki değer alınır
                                fonksiyonsuz_kalan_zaman = df_selected.iloc[2, int_column_fonksiyonsuz]
                                tablet_kalan_zaman = df_selected.iloc[2, int_column_tablet]

                                if tablet_kalan_zaman > fonksiyonsuz_kalan_zaman:
                                    pass
                                else:
                                    current_file.error_elements10.append("HATA")

                            # ----------------------------------------------------------
                            # 11. Kural - Sirkülasyon/Tahliye adımları 15 sn den düşük olmalıdır (BLDC) , Asenkorn için 25 sn

                            # SIRKULASYON / TAHLIYE içeren satırlara göre filtreleme

                            file_name = self.excel_path.split("/")[-1]
                            name_parts = file_name.split("_")

                            condition_sirkulasyon = df_selected.iloc[:, 2] == "SIRKULASYON / TAHLIYE"
                            filtered_df = df_selected[condition_sirkulasyon]
                            # satırlardaki sn değişkenlerini değerlendirme
                            for row_index in range(len(filtered_df)):
                                for col_index in range(1, len(filtered_df.columns)):
                                    if filtered_df.iloc[row_index, col_index] == 'sn':
                                        degisken = filtered_df.iloc[row_index, col_index - 1]
                                        asenkron = False
                                        name_digit = int(name_parts[0][6])
                                        if name_digit == 1:
                                            asenkron = True
                                        if asenkron:
                                            if degisken < 0 or degisken > 25:
                                                degisken_col = self.column_number_to_letter(col_index - 1)
                                                current_file.error_elements11.append(
                                                    f'{filtered_df.index[row_index] + 2}{degisken_col}')
                                        else:
                                            if degisken < 0 or degisken > 15:
                                                degisken_col = self.column_number_to_letter(col_index - 1)
                                                current_file.error_elements11.append(
                                                    f'{filtered_df.index[row_index] + 2}{degisken_col}')

                            # ----------------------------------------------------------
                            # 12.Kural - Akışların içinde hijyen fonksiyonu olmalı ve tablet fonksiyonu olmalıdır (makine temizleme ve ön yıkama akışları hijyen ve tablet içermez)
                            file_name = self.excel_path.split("/")[-1]
                            check_makine_temizleme = "MakineTemizleme" in file_name
                            check_pre_wash = "PreWash" in file_name
                            check_mini30 = "Mini30" in file_name

                            check_tablet = False
                            check_hygiene = False

                            if not (check_makine_temizleme or check_pre_wash):
                                if not check_mini30:
                                    # fonksiyonlar HYGIENE ve TABLET fonksiyonu içeriyor mu kontrolü
                                    for col in df_selected.iloc[0, :]:
                                        if "TABLET" in str(col):
                                            check_tablet = True
                                        elif "HYGIENE" in str(col):
                                            check_hygiene = True
                                    if not check_tablet:
                                        current_file.error_elements12.append(
                                            "TABLET fonk. bulunmuyor.")
                                    if not check_hygiene:
                                        current_file.error_elements12.append(
                                            "HYGIENE fonk. bulunmuyor.")
                                else:
                                    # mini30 akışları için sadece tablet kontrolü
                                    for col in df_selected.iloc[0, :]:
                                        if "TABLET" in str(col):
                                            check_tablet = True
                                            break
                                    if not check_tablet:
                                        current_file.error_elements12.append(
                                            "TABLET fonk. bulunmuyor.")

                            # ----------------------------------------------------------
                            # 13.Kural - TABLET fonksiyonunda 2.SCD'de C olan adımlar arasında toplam sirkülasyon min 4.5dk max 8dk olmalıdır

                            if not check_tablet:

                                # tablet bloğunu istenilen koşullarla oluşturma
                                tablet_block = sicakDurulama2_block.iloc[:,
                                               [2, 3, int_column_tablet, int_column_tablet + 1, int_column_tablet + 2]]

                                # sirkülasyon süresi kontrolü
                                check_sicaklik = False
                                first_row = -1
                                second_row = -1
                                toplam = 0
                                for row in range(0, len(tablet_block)):
                                    # SICAKLIK adımları
                                    if tablet_block.iloc[row, 1] == "SICAKLIK":
                                        if not check_sicaklik:
                                            # Starting a new block
                                            check_sicaklik = True
                                            first_row = sicakDurulama2_rowNumber + row
                                            toplam = 0  # Reset toplam for the new block
                                        else:
                                            # Ending the current block
                                            check_sicaklik = False
                                            second_row = sicakDurulama2_rowNumber + row
                                            # Total circulation time check
                                            if not (4.5 <= toplam <= 8):
                                                current_file.error_elements13.append(
                                                    f"{first_row} ve {second_row} arasında toplam sirkülasyon zamanı hatası")

                                            # Starting a new block immediately
                                            check_sicaklik = True
                                            first_row = sicakDurulama2_rowNumber + row
                                            toplam = 0  # Reset toplam for the new block
                                    # SIRKULASYON adımları
                                    elif check_sicaklik and "SIRKULASYON" in tablet_block.iloc[row, 0]:
                                        if tablet_block.iloc[row, 1] == "SURE" and tablet_block.iloc[row, 4] == "dak":
                                            toplam += tablet_block.iloc[row, 3]
                                        elif tablet_block.iloc[row, 1] == "SURE" and tablet_block.iloc[row, 4] == "sn":
                                            toplam += tablet_block.iloc[row, 3] / 60

                            # ----------------------------------------------------------
                            # # 14.Kural - Adım sayısı arttıkça adıma ait süre değeri artamaz
                            #
                            # # bütün data frame'den kalan zaman blokları parçasını alma
                            # kalan_zaman_blok = df_selected.iloc[2:stop_rowNumber - 2, 4::4]
                            #
                            # # kalan zaman değerlerinin aşağıya doğru azalması gerekiyor
                            # check = True
                            # for col in range(0, len(kalan_zaman_blok.columns)):
                            #     min_element = kalan_zaman_blok.iloc[0, col]
                            #     for row in range(1, len(kalan_zaman_blok)):
                            #         if min_element >= kalan_zaman_blok.iloc[row, col]:
                            #             min_element = kalan_zaman_blok.iloc[row, col]
                            #         else:
                            #             if not row == len(kalan_zaman_blok) - 1:
                            #                 check = False
                            #                 # hatalı sütunu hesaplama
                            #                 if col == 0:
                            #                     error_col = self.column_number_to_letter(4)
                            #                 else:
                            #                     error_col = self.column_number_to_letter((col + 1) * 4)
                            #                 self.error_elements14.append(f"{row + 4}{error_col}")
                            #                 break
                            #
                            # if not check:
                            #     current_file.error_elements14.append("HATA")

                            # ----------------------------------------------------------
                            # 15.Kural - Fonksiyon sütunu ısıtıcı ise ve kontrol sütunu süre ise ve fonksiyon sütünü dakika ise 15 dakikadan fazla olmamalı

                            # aranılan koşullara göre
                            contains_ısıtıcı = df_selected.iloc[:, 2].str.contains("ISITICI")
                            condition_sure = df_selected.iloc[:, 3] == "SURE"
                            selected_rows = condition_sure & contains_ısıtıcı
                            indexes = []
                            for i in range(0, len(selected_rows)):
                                if selected_rows.iloc[i]:
                                    indexes.append(i)
                            for row in indexes:
                                for col in range(6, len(df_selected.columns), 4):
                                    if df_selected.iloc[row, col] == "dak" and df_selected.iloc[row, col - 1] > 15:
                                        current_file.error_elements15.append(
                                            f"{row + 2}{self.column_number_to_letter(col - 1)}")

                            # ----------------------------------------------------------
                            # 16.Kural - Rej'li sütunlar hariç 2.SCD adımında Parlatıcı adımında süre 80 sn den büyük eşit olmalı

                            # DETERJAN satırlarını bulma
                            condition_deterjan = sicakDurulama2_block.iloc[:, 2].str.contains("DETERJAN")
                            c_sure = sicakDurulama2_block.iloc[:, 3] == "SURE"
                            scd2_deterjan_rows = sicakDurulama2_block[condition_deterjan & c_sure]
                            # satırların sn değerlerini kontrol etme
                            for row in range(0, len(scd2_deterjan_rows)):
                                for col in range(6, len(scd2_deterjan_rows.columns), 4):
                                    # eğer rej sütunlarında değilse
                                    if not ("Rej" in df_selected.iloc[0, col - 2]):
                                        if scd2_deterjan_rows.iloc[row, col] == "sn":
                                            if scd2_deterjan_rows.iloc[row, col - 1] < 80:
                                                current_file.error_elements16.append(
                                                    f"{scd2_deterjan_rows.index[row] + 2}{self.column_number_to_letter(col - 1)}")

                            # ----------------------------------------------------------
                            # 17.Kural - Kapı açma için maksimum süre 110sn olmalıdır

                            # KAPI ACMA satırlarını bulma
                            condition_kapı_acma = df_selected.iloc[:, 2].str.contains("KAPI ACMA")
                            # condition_sure = df_selected.iloc[:, 3] == "SURE" (yukarıda tanımlı)
                            kapı_acma_rows = df_selected[condition_kapı_acma & condition_sure]

                            # KAPI ACMA sn/dak değerleri kontrol
                            for row in range(0, len(kapı_acma_rows)):
                                for col in range(6, len(kapı_acma_rows.columns), 4):
                                    if kapı_acma_rows.iloc[row, col] == "sn" and kapı_acma_rows.iloc[
                                        row, col - 1] > 110:
                                        current_file.error_elements17.append(
                                            f"{kapı_acma_rows.index[row] + 2}{self.column_number_to_letter(col - 1)}")
                                    elif kapı_acma_rows.iloc[row, col] == "dak" and kapı_acma_rows.iloc[
                                        row, col - 1] > 1.83:
                                        current_file.error_elements17.append(
                                            f"{kapı_acma_rows.index[row] + 2}{self.column_number_to_letter(col - 1)}")

                            # ----------------------------------------------------------
                            # 18.Kural - Kurutma bloğunun BosaltmaVanası/Tahliye adımı ile Su giriş vanası arasında Fan,bekleme veya Fan/klape adımlarının olması gerekmektedir

                            # aranan satırların indexleri bulma
                            first_row = -1
                            last_row = -1
                            wanted_row = -1
                            for row in range(0, len(kurutma_block)):
                                element = kurutma_block.iloc[row, 2]
                                if first_row == -1 and (
                                        element == "TAHLIYE" or element == "BOSALTMA VANASI" or element == "BOSALTMA VANASI / TAHLIYE"):
                                    first_row = row
                                elif wanted_row == -1 and (
                                        element == "FAN" or element == "BEKLEME" or element == "FAN / KLAPE"):
                                    wanted_row = row
                                elif last_row == -1 and element == "SU GIRIS VANASI":
                                    last_row = row
                            # bulunan indexler istenileni sağlıyor mu kontrolü
                            check = False
                            if first_row != -1 and wanted_row != -1 and last_row != -1:
                                if first_row < wanted_row < last_row:
                                    check = True

                            if not check:
                                current_file.error_elements18.append("HATA")

                            # ----------------------------------------------------------
                            # 19.Kural - Vanalı akışlarda Parlatıcı adımında vana pozisyonu kapalı (idle) olmamalı
                            file_name = self.excel_path.split("/")[-1]
                            name_parts = file_name.split("_")
                            # vanalı için dosya isminin baştan 6. karakteri kontrol edilir. (0 olmamalı)
                            value = int(name_parts[0][5])
                            vanali = value not in [0, 4]

                            #kv1000 için dosya isminin baştan 6. karakteri kontrol edilir. (8 olmalı)
                            kv1000_check = int(name_parts[0][5]) == 8

                            if kv1000_check:
                                valid_vana_positions = ["ALT", "UST", "ZONE", "RUN", "KAPALI"]
                            else:
                                valid_vana_positions = ["ALT", "UST", "TEPE", "ZONE", "RUN", "KAPALI"]

                            # DETERJAN içeren satırları bulma
                            c_deterjan = df_selected.iloc[:, 2].str.contains("DETERJAN")
                            deterjan_rows = df_selected[c_deterjan & 1]

                            # akış vanalı ise kural kontrol edilir
                            if vanali:
                                for row in range(0, len(deterjan_rows)):
                                    for col in range(7, len(deterjan_rows.columns), 4):
                                        vana_values = deterjan_rows.iloc[row, col].split("-")
                                        # vana_values[1, 3, 5, 7] kontrol edilecek
                                        if deterjan_rows.iloc[row, col - 2] != 0:
                                            if not (vana_values[1] in valid_vana_positions and vana_values[3] in valid_vana_positions and
                                                    vana_values[5] in valid_vana_positions and vana_values[7] in valid_vana_positions):
                                                current_file.error_elements19.append(
                                                    f"{deterjan_rows.index[row] + 2}{self.column_number_to_letter(col)}")

                            # -----------------------------------------------------------
                            # "20.Kural - Tahliye adımı sayısı su alma adımı sayısından en az bir fazla olmalıdır. "
                            # "(Bir fazla olma nedeni programın tahliye ile başlaması)\n"
                            # "    a. Tahliye adımları ile su alma adımlarının sayısı uyuşmuyorsa "
                            # "\"Fonksiyon içerisinde tahliye adımlarında dengesizlik bulunmakta, kontrol edin\" "
                            # "uyarısı verilmekte.\n"
                            # "    b. Her Tahliye bloğunun ardından su girişi olmalıdır. Son tahliye hariç."

                            # fonksiyonlardaki dolu tahliye ve su giriş vanası adımlarının sayılarını ve sırasını tutma

                            non_rej_cols = []
                            for col in range(4, len(df_selected.columns), 4):
                                if "(Rej)" not in df_selected.iloc[0, col] and not pd.isna(df_selected.iloc[0, col]):
                                    non_rej_cols.append(col + 1)

                            # son TAHLIYE satırının indexini bulma
                            last_tahliye_row = -1
                            for row in range(0, len(df_selected)):
                                if df_selected.iloc[row, 2] == "TAHLIYE":
                                    last_tahliye_row = row

                            fonksiyon_counts = {}
                            fonksiyon_order = {}
                            for col in non_rej_cols:
                                fonksiyon_counts[col] = 0
                                fonksiyon_order[col] = ""

                            for row in range(2, stop_rowNumber - 2):
                                if df_selected.iloc[row, 2] == "TAHLIYE":
                                    for col in non_rej_cols:
                                        if df_selected.iloc[row, col] > 0:
                                            fonksiyon_counts[col] += 1
                                            if row != last_tahliye_row:
                                                fonksiyon_order[col] += "T"
                                elif df_selected.iloc[row, 2] == "SU GIRIS VANASI":
                                    for col in non_rej_cols:
                                        if df_selected.iloc[row, col] > 0:
                                            fonksiyon_counts[col] -= 1
                                            fonksiyon_order[col] += "S"
                            print(fonksiyon_order)
                            # fonksiyonlardaki dolu tahliye ve su giriş vanası adımlarının sayılarını ve sırasını kontrol
                            for index in fonksiyon_counts.keys():
                                if fonksiyon_counts[index] < 1:
                                    current_file.error_elements20.append(
                                        f"{df_selected.iloc[0, index - 1]} içerisinde tahliye adımlarında dengesizlik bulunmakta")
                                if "TT" in fonksiyon_order[index]:
                                    current_file.error_elements20.append(
                                        f"{df_selected.iloc[0, index - 1]} içerisinde tahliye/su giriş sırasında dengesizlik bulunmakta")

                            # ----------------------------------------------------------
                                    # 21.Kural - Parlatıcı atılan içinde deterjan geçen adım süresi 80 sn den düşük olmamalı

                                    # deterjan_rows yukarıda tanımlı
                                    # DETERJAN satırlarında sn değerlerini kontrol
                                    for row in range(0, len(deterjan_rows)):
                                        if deterjan_rows.iloc[row, 3] == "SURE":
                                            for col in range(6, len(deterjan_rows.columns), 4):
                                                if deterjan_rows.iloc[row, col] == "sn" and deterjan_rows.iloc[
                                                    row, col - 1] < 80 and deterjan_rows.iloc[row, col - 1] != 0:
                                                    current_file.error_elements21.append(
                                                        f"{deterjan_rows.index[row] + 2}{self.column_number_to_letter(col - 1)}")
                                                elif deterjan_rows.iloc[row, col] == "dak" and deterjan_rows.iloc[
                                                    row, col - 1] < 1.33 and deterjan_rows.iloc[row, col - 1] != 0:
                                                    current_file.error_elements21.append(
                                                        f"{deterjan_rows.index[row] + 2}{self.column_number_to_letter(col - 1)}")

                            # ----------------------------------------------------------
                            # 22.Kural

                            # değerlendirilen adım sayısı olarak dolu olan adım sayılarını kabul ettim
                            # dolu adımların sayısını hesaplama

                            if not "ODD" in current_file.file_name:

                                deterjan_rows = df_selected[df_selected.iloc[:, 2].eq("SIRKULASYON / ISITICI / DETERJAN") & 1]

                                # non_rej_cols 20.kuralda tanımlı
                                fonksiyon_step_count = {}
                                for col in non_rej_cols:
                                    fonksiyon_step_count[col] = 0

                                for row in range(0, len(deterjan_rows)):
                                    for col in non_rej_cols:
                                        if deterjan_rows.iloc[row, col] > 0:
                                            fonksiyon_step_count[col] += 1

                                # hangi fonksiyonlarda kaç adım var kontrolü ve hatalı fonksiyonları bulma
                                for col, value in fonksiyon_step_count.items():
                                    if value > 2:
                                        current_file.error_elements22.append(
                                            f"{df_selected.iloc[0, col-1]} Fonksiyonu içerisinde fazla deterjan adımı bulunmakta, kontrol edin")
                                    elif value < 1:
                                        current_file.error_elements22.append(
                                            f"{df_selected.iloc[0, col-1]} Fonksiyonu içerisinde az deterjan adımı bulunmakta, kontrol edin")

                            # ----------------------------------------------------------
                            # 23.Kural - MSC2 akışların su alma adımlarından sonra boşaltma vanası ve bekleme olmalı ** iki su giriş arasında olsun

                            file_name = self.excel_path.split("/")[-1]  # Dosya yolundan dosya adını al
                            # dosyanın adı ve boşaltma vanası ile MSC2 mi değil mi kontrolü
                            check_bosaltma = False
                            for row in range(sicakDurulama2_rowNumber + sicakDurulama2_adim_sayisi - 2, len(df_selected)):

                                if "BOSALTMA VANASI" in df_selected.iloc[row, 2]:
                                    if 0 in df_selected.iloc[row, 5:: 4].values:
                                        check_bosaltma = True
                                        break
                                else:
                                    print("boşaltma yok")
                                    break

                            msc2_check = check_bosaltma or "MSC2" in file_name
                            if msc2_check:
                                check_su = False
                                check_bek = False
                                check_bos = False
                                main_check = False
                                # istenilen satırları arama
                                for row in range(0, len(df_selected)):
                                    if df_selected.iloc[row, 2] == "SU GIRIS VANASI":
                                        if not check_su:
                                            check_su = True
                                        else:
                                            break
                                    if check_su and df_selected.iloc[row, 2] == "BEKLEME":
                                        check_bek = True
                                    if check_su and df_selected.iloc[row, 2] == "BOSALTMA VANASI":
                                        check_bos = True
                                    if check_su and check_bek and check_bos:
                                        check_su = False
                                        check_bek = False
                                        check_bos = False
                                if not (check_su or check_bos or check_bek):
                                    main_check = True
                            else:
                                main_check = False

                            if not main_check and msc2_check:
                                current_file.error_elements23.append("HATA")

                            # ----------------------------------------------------------
                            # 24.Kural (sadece kapı açmali sütunlarda bakılacak)

                            # Mini30 ve Hizli58 için excel isimleri kontrol edilir
                            kapı_acma_check = False
                            file_name = self.excel_path.split("/")[-1]
                            if "Hizli58" in current_file.file_name or "Mini30" in current_file.file_name :
                                # kapı açma sütun indexlerini bulma
                                kapı_acma_cols = []
                                for col in range(4, len(df_selected.columns), 4):
                                    if "KAPI AÇMA" in str(df_selected.iloc[0, col]):
                                        kapı_acma_cols.append(col)
                                fonksiyon_dereceleri = [0] * len(kapı_acma_cols)
                                # kapı açma sütunlarının son sıcaklık değerlerini bulma
                                for row in range(0, len(sicakDurulama2_block)):
                                    if sicakDurulama2_block.iloc[row, 3] == "SICAKLIK":
                                        if sicakDurulama2_block.iloc[row, 6] == "C":
                                            for index in range(0, len(kapı_acma_cols)):
                                                fonksiyon_dereceleri[index] = sicakDurulama2_block.iloc[
                                                    row, kapı_acma_cols[index] + 1]

                                # kapı açma sütunları için kurutma kısmındaki toplam bekleme süreleri hesaplama
                                toplam_bekleme = [0] * len(fonksiyon_dereceleri)
                                for row in range(0, len(kurutma_block)):
                                    if kurutma_block.iloc[row, 2] == "BEKLEME" or kurutma_block.iloc[row, 2] == "FAN" \
                                            or kurutma_block.iloc[row, 2] == "FAN / KLAPE":
                                        for index in range(0, len(kapı_acma_cols)):
                                            if kurutma_block.iloc[row, kapı_acma_cols[index] + 2] == "dak":
                                                toplam_bekleme[index] += kurutma_block.iloc[
                                                    row, kapı_acma_cols[index] + 1]
                                            elif kurutma_block.iloc[row, kapı_acma_cols[index] + 2] == "sn":
                                                toplam_bekleme[index] += kurutma_block.iloc[
                                                                             row, kapı_acma_cols[index] + 1] / 60
                                # sütunların toplam bekleme süreleri ile derecelerini karşılaştırma
                                check = True
                                for index in range(0, len(fonksiyon_dereceleri)):
                                    if fonksiyon_dereceleri[index] > 50:
                                        min_sure = fonksiyon_dereceleri[index] - 50
                                        if toplam_bekleme[index] < min_sure:
                                            current_file.error_elements24.append(
                                                f"{df_selected.iloc[0, 4 * index + 4]} HATA")
                                            kapı_acma_check = True

                            if check:
                                pass
                            else:
                                if kapı_acma_check:
                                    current_file.error_elements24.append("KAPI AÇMA HATA")
                                else:
                                    print("Kapı açma yok")

                            # ----------------------------------------------------------
                            # 25.Kural - İki su alma adımı arasındaki toplam tahliye süresi 30 sn'den fazla 60sn'den az olmalıdır

                            # # su giriş vanası olan satırların indexlerini bulma
                            # su_giris_rows = []
                            # for row in range(0, len(df_selected)):
                            #     if df_selected.iloc[row, 2] == "SU GIRIS VANASI":
                            #         su_giris_rows.append(row)
                            # # iki su giriş indexi arasına bakılarak her fonksiyon için ayrı ayrı tahliye adımlarını toplama
                            # check = True
                            # for index in range(0, len(su_giris_rows) - 1):
                            #     toplam = [0] * (((len(df_selected.columns) - 5) // 4) + 1)
                            #     for i in range(su_giris_rows[index], su_giris_rows[index + 1]):
                            #         if df_selected.iloc[i, 2] == "TAHLIYE" and df_selected.iloc[i, 3] == "SURE":
                            #             for col in range(5, len(df_selected.columns), 4):
                            #                 if df_selected.iloc[i, col + 1] == "sn":
                            #                     toplam[(col - 5) // 4] += df_selected.iloc[i, col]
                            #                 elif df_selected.iloc[i, col + 1] == "dak":
                            #                     toplam[(col - 5) // 4] += df_selected.iloc[i, col] * 60
                            #     for top in range(0, len(toplam)):
                            #         if not (30 < toplam[top] < 60):
                            #             if not current_file.error_elements25.__contains__(f"{df_selected.iloc[0, 4 * index + 4]} Fonksiyonu içerisinde HATA"):
                            #                 current_file.error_elements25.append(
                            #                     f"{df_selected.iloc[0, 4 * index + 4]} Fonksiyonu içerisinde HATA")

                            # ----------------------------------------------------------
                            # 26.Kural - Akış isimleri EXTENDED sayfasına ve diğer sayfalara doğru girilmeli. (excelin isim sheetlere doğru girilmeli)

                            # isim ve extended sayfasında kontrol edilecek kısımları belirleme
                            name_parts = file_name.split("_")
                            print(name_parts)
                            extended_parts = excel_file.sheet_names[extended_index[0]].split("_")

                            if name_parts[0] == (extended_parts[0]):
                                pass
                            else:
                                current_file.error_elements26.append("HATA")

                            # ----------------------------------------------------------
                            # 27.Kural - Janus tahliyeler için RPM değerleri girilmelidir.
                            # janus için dosya isminin baştan 7. karakteri kontrol edilir. (5 veya 7 olması gerek)
                            janus = False
                            if int(name_parts[0][6]) == 5 or int(name_parts[0][6]) == 7:
                                janus = True
                            if janus:
                                # rpm değerlerini kontrol
                                for row in range(2, stop_rowNumber - 2):
                                    if df_selected.iloc[row, 2] == "TAHLIYE":
                                        for col in range(7, len(df_selected.columns), 4):
                                            cvv_cells = df_selected.iloc[row, col].split("/")
                                            flat_list = [item for part in cvv_cells for item in part.split('-')]
                                            result = [int(item.strip()) if item.strip().isdigit() else item.strip() for
                                                      item in
                                                      flat_list]
                                            if result[2] == 0 or result[5] == 0 or result[8] == 0 or result[11] == 0:
                                                current_file.error_elements27.append(
                                                    f"{row + 2}{self.column_number_to_letter(col)}")

                            # ----------------------------------------------------------
                            # 28.Kural - Ana yıkama , 1SCD , 2.SCD bloklarında en az 1 tane "C, dak, sn" yazan ısıtıcı adımları 0 dan farklı olacak

                            # Ana Yıkama bloğu kontrol
                            ana_yikama_check = False
                            error_cells = []
                            for row in range(0, len(anaYikama_block)):
                                if "ISITICI" in anaYikama_block.iloc[row, 2] and anaYikama_block.iloc[
                                    row, 3] == "SICAKLIK":
                                    for col in range(5, len(anaYikama_block.columns), 4):
                                        birim = anaYikama_block.iloc[row, col + 1]
                                        if birim == "C" or birim == "dak" or birim == "sn":
                                            if anaYikama_block.iloc[row, col] <= 0:
                                                error_cells.append(anaYikama_block.iloc[row, col])
                                            else:
                                                error_cells.append(anaYikama_block.iloc[row, col])
                                                ana_yikama_check = True

                            print(f"ANA YIKAMA: {sum(error_cells)}")
                            error_cells = []

                            # 1.SCD bloğu kontrol
                            scd_1_check = False
                            for row in range(0, len(sicakDurulama1_block)):
                                if "ISITICI" in sicakDurulama1_block.iloc[row, 2] and sicakDurulama1_block.iloc[
                                    row, 3] == "SICAKLIK":
                                    for col in range(5, len(sicakDurulama1_block.columns), 4):
                                        birim = sicakDurulama1_block.iloc[row, col + 1]
                                        if birim == "C" or birim == "dak" or birim == "sn":
                                            if sicakDurulama1_block.iloc[row, col] <= 0:
                                                error_cells.append(sicakDurulama1_block.iloc[row, col])
                                            else:
                                                error_cells.append(sicakDurulama1_block.iloc[row, col])
                                                scd_1_check = True

                            print(f"1.SCD: {sum(error_cells)}")
                            error_cells = []

                            # 2.SCD bloğu kontrol
                            scd_2_check = False
                            for row in range(0, len(sicakDurulama2_block)):
                                if "ISITICI" in sicakDurulama2_block.iloc[row, 2] and sicakDurulama2_block.iloc[
                                    row, 3] == "SICAKLIK":
                                    for col in range(5, len(sicakDurulama2_block.columns), 4):
                                        birim = sicakDurulama2_block.iloc[row, col + 1]
                                        if birim == "C" or birim == "dak" or birim == "sn":
                                            if sicakDurulama2_block.iloc[row, col] <= 0:
                                                error_cells.append(sicakDurulama2_block.iloc[row, col])
                                        else:
                                            error_cells.append(sicakDurulama2_block.iloc[row, col])
                                            scd_2_check = True

                            print(f"2.SCD: {sum(error_cells)}")

                            if not ana_yikama_check:
                                current_file.error_elements28.append("ANA YIKAMA BLOĞU HATA")
                            if not scd_1_check:
                                current_file.error_elements28.append("1.SCD BLOĞU HATA")
                            if not scd_2_check:
                                current_file.error_elements28.append("2.SCD BLOĞU HATA")

                            # ----------------------------------------------------------
                            # 29.Kural - Asenkron Tahliyeler için RPM değeri 0 olmalıdır

                            # asenkron için dosya isminin baştan 7. karakteri kontrol edilir. (0, 1, 2, 3, 4 olması gerek)
                            asenkron = False
                            name_digit = int(name_parts[0][6])
                            if 0 <= name_digit <= 4:
                                asenkron = True
                            if asenkron:
                                # rpm değerlerini kontrol
                                for row in range(2, stop_rowNumber - 2):
                                    if df_selected.iloc[row, 2] == "TAHLIYE":
                                        for col in range(7, len(df_selected.columns), 4):
                                            cvv_cells = df_selected.iloc[row, col].split("/")
                                            flat_list = [item for part in cvv_cells for item in part.split('-')]
                                            result = [int(item.strip()) if item.strip().isdigit() else item.strip() for
                                                      item in flat_list]
                                            if not (result[2] == 0 and result[5] == 0 and result[8] == 0 and result[
                                                11] == 0):
                                                current_file.error_elements29.append(
                                                    f"{row + 2}{self.column_number_to_letter(col)}")

                            # ----------------------------------------------------------
                            # 30.Kural - Vanalılarda vanalı RPM ve pozisyon değerlerinin girilmiş olması gerekmektedir. (su giris vanası ve tahliye dışındakilere bak)

                            # akış vanalı ise kural kontrol edilir
                            # vanali 19.kuralda tanımlı
                            if vanali and not asenkron:
                                # rpm ve vana değerleri kontrol
                                # valid_vana_positions = ["ALT", "UST", "TEPE", "ZONE", "RUN", "KAPALI"]  yukarıda tanımlı
                                for row in range(2, stop_rowNumber - 2):
                                    if not (df_selected.iloc[row, 2] == "TAHLIYE" or df_selected.iloc[row, 2] == "SU GIRIS VANASI"
                                            or "SIRKULASYON" in df_selected.iloc[row, 2]):
                                        for col in range(7, len(df_selected.columns), 4):
                                            if df_selected.iloc[row, col - 2] != 0:
                                                cvv_cells = df_selected.iloc[row, col].split("/")
                                                flat_list = [item for part in cvv_cells for item in part.split('-')]
                                                result = [int(item.strip()) if item.strip().isdigit() else item.strip()
                                                          for item in flat_list]
                                                condition1 = not (result[1] in valid_vana_positions and result[
                                                    4] in valid_vana_positions and
                                                                  result[7] in valid_vana_positions and result[
                                                                      10] in valid_vana_positions)
                                                condition2 = result[2] == 0 or result[5] == 0 or result[8] == 0 or \
                                                             result[11] == 0
                                                if condition1 or condition2:
                                                    current_file.error_elements30.append(
                                                        f"{row + 2}{self.column_number_to_letter(col)}")

                            # ----------------------------------------------------------
                            # 31.Kural - Strong ve Strong+ kombinasyonlarda SCD1 bloğunda boşaltma vanası olmalıdır.

                            # STRONG içeren sütunların indexlerini bulmak
                            strong_indices = []
                            for col in df_selected.columns:
                                if col.__contains__("Unnamed"):
                                    element = str(df_selected.iloc[0, int(col[9:])])
                                    if "STRONG" in element:
                                        strong_indices.append(int(col[9:]))
                            bosaltma_rows = []
                            # STRONG kombinasyonları için SCD1 kontrolü
                            if len(strong_indices) != 0:
                                for row in range(0, len(sicakDurulama1_block)):
                                    if sicakDurulama1_block.iloc[row, 2] == "BOSALTMA VANASI":
                                        bosaltma_rows.append(row)
                                        for col in strong_indices:
                                            if sicakDurulama1_block.iloc[row, col + 1] <= 0:
                                                current_file.error_elements31.append(
                                                    f"{sicakDurulama1_block.index[row] + 2}{self.column_number_to_letter(col + 1)}")

                            # ----------------------------------------------------------
                            # 32.Kural - Auto programlarda turbitity ölçüm adımları kontrol edilmeli dolu olması gerekmektedir. (0 kir ve yarım kir hariç)

                            # dosya ismi AUTO içeriyor mu kontrolü
                            # auto = False
                            # if "Auto" in current_file.file_name:
                            #     auto = True
                            #
                            # if isinstance(col, str):
                            #     try:
                            #         col = int(col)
                            #     except ValueError:
                            #         print(f"col'u tam sayıya dönüştürürken hata: {col}")
                            #         continue  # veya hatayı uygun şekilde yönetin
                            #
                            #
                            # # turbitity satırlarını kontrol
                            # if auto:
                            #     for row in range(2, stop_rowNumber - 2):
                            #         if not (("Rej" in df_selected.iloc[0, col - 2]) or ("Yarım Kir" in df_selected.iloc[0, col - 2]) or ("0 Kir" in df_selected.iloc[0, col - 2])) :
                            #             if " TURBIDITY" in df_selected.iloc[row, 2]:
                            #                 for col in range(5, len(df_selected.columns), 4):
                            #                     if df_selected.iloc[row, col] <= 0:
                            #                         current_file.error_elements32.append(f"{row + 2}{self.column_number_to_letter(col)}")
                            # else:
                            #     print("auto yok")

                            # ----------------------------------------------------------
                            # 33.Kural - Rejenerasyon ve Su giriş vanası sürülüyorsa tahliye olmalıdır. (RY Bloğundandan sonraki adımda tahliye olmalı)
                            # !!!!!! REJENERASYON / SU GIRIS VANASI adımı 0sa o sütunun tahliyesini kontrol etmedim !!!!!!

                            # (Rej) içeren sütunların indexlerini bulma
                            rej_indices = []
                            for col in df_selected.columns:
                                if col.__contains__("Unnamed"):
                                    element = str(df_selected.iloc[0, int(col[9:])])
                                    if "(Rej)" in element:
                                        rej_indices.append(int(col[9:]))

                            # RY bloğundaki satırları bulma
                            for row in range(0, len(rejenarasyon_block)):
                                if "REJENERASYON / SU GIRIS VANASI" in rejenarasyon_block.iloc[row, 2]:
                                    for col in rej_indices:
                                        if rejenarasyon_block.iloc[row, col + 1] <= 0:
                                            rej_indices.remove(col)

                            # tahliye adımını kontrol
                            for row in range(rejenarasyon_rowNumber + rejenarasyon_adim_sayisi - 2, len(df_selected)):
                                if "TAHLIYE" in str(df_selected.iloc[row, 2]):
                                    for col in rej_indices:
                                        if df_selected.iloc[row, col + 1] <= 0:
                                            current_file.error_elements33.append(
                                                f"{row + 2}{self.column_number_to_letter(col + 1)}")
                                    break

                            # ----------------------------------------------------------
                            # 34.Kural - Rejenerasyonlu fonksiyonlarda rejenerasyon adımının dolu olması

                            # excelin sütunlarını okuyarak fonksiyonların (Rej)'li halini kontrol etme
                            check = True
                            fonksiyonlar = df_selected.iloc[0, 4:: 4]
                            #fonksiyonlar = [str(fonk) for fonk in df_selected.iloc[0, 4::4]]
                            for fonksiyon in fonksiyonlar:
                                if "(Rej)" not in fonksiyon:
                                    modified_fonksiyon = fonksiyon.split('(')[0].strip()
                                    fonksiyon_rej = next(
                                        (item for item in fonksiyonlar if f"{modified_fonksiyon} (Rej)" in item), None)
                                    if fonksiyon_rej is None:
                                        current_file.error_elements34.append(fonksiyon)
                                else:
                                    break

                            # ----------------------------------------------------------
                            # 35.Kural - ½ fonksiyonunun su tüketimi ve süresi fonksiyonsuzdan az olmalı

                            # FONKSİYONSUZ ve HALF LOAD tanımlı mı kontrolü
                            if check_fonksiyonsuz and check_halfLoad:
                                int_column_half_load = int(column_halfLoad[9:])
                                # kalan zaman kontrolü
                                check_kalan_zaman = False
                                if df_selected.iloc[2, int_column_half_load] < df_selected.iloc[
                                    2, int_column_fonksiyonsuz]:
                                    check_kalan_zaman = True

                                # su tüketimi kontrolü (SU GIRIS VANASI satırlarındaki değerlerin toplamlarının karşılaştırması)
                                half_load_toplam = 0
                                fonksiyonsuz_toplam = 0
                                for row in range(2, stop_rowNumber - 2):
                                    if df_selected.iloc[row, 2] == "SU GIRIS VANASI" and df_selected.iloc[
                                        row, 3] == "SU MIKTARI":
                                        half_load_toplam += df_selected.iloc[row, int_column_half_load + 1]
                                        fonksiyonsuz_toplam += df_selected.iloc[row, int_column_fonksiyonsuz + 1]

                                if not check_kalan_zaman:
                                    current_file.error_elements35.append("Süre Hatası")
                                if half_load_toplam > fonksiyonsuz_toplam:
                                    current_file.error_elements35.append("Toplam Su Tüketimi Hatası")

                            # ----------------------------------------------------------
                            # 36.Kural - Fast süresinin fonksiyonsuzdan kısa olması ve ana yıkama sıcaklığının fonksiyonsuzdan yüksek olmalı

                            if check_fast and check_fonksiyonsuz:
                                int_column_fast = int(column_fast[9:])
                                # kalan zaman kontrolü
                                check_kalan_zaman = False
                                if df_selected.iloc[2, int_column_fast] < df_selected.iloc[2, int_column_fonksiyonsuz]:
                                    check_kalan_zaman = True

                                # ANA YIKAMA bloğundaki sıcaklık değerlerinin toplamını bulma
                                fast_toplam = 0
                                fonksiyonsuz_toplam = 0
                                for row in range(0, len(anaYikama_block)):
                                    if anaYikama_block.iloc[row, 3] == "SICAKLIK":
                                        fast_toplam += anaYikama_block.iloc[row, int_column_fast + 1]
                                        fonksiyonsuz_toplam += anaYikama_block.iloc[row, int_column_fonksiyonsuz + 1]

                                if not check_kalan_zaman:
                                    current_file.error_elements36.append("Süre Hatası")
                                elif fast_toplam < fonksiyonsuz_toplam:
                                    current_file.error_elements36.append("Ana Yıkama Sıcaklık Hatası")

                            # ----------------------------------------------------------
                            # 37.Kural - 60cm Vanalı ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 202 dakika olmalıdır.

                            # vanali 19.kuralda tanımlı
                            # 60cm ve Yoğun 70 dosya isminden bakılacak
                            # 60cm genişlik için dosya isminin 4. karakterine bakılır (3 olmalı)

                            genislik_value = int(name_parts[0][3])
                            hygiene_col = -1
                            # hygiene sütununun indexini bulma
                            for col in range(0, len(df_selected.columns)):
                                if "HYGIENE" in str(df_selected.iloc[0, col]):
                                    hygiene_col = col
                                    break
                            print(vanali)
                            if vanali and genislik_value == 3 and "Intensive70" in current_file.file_name and not hygiene_col == -1:
                                # hygiene sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, hygiene_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, hygiene_col]))
                                        if not df_selected.iloc[row + 1, hygiene_col] == 202:
                                            current_file.error_elements37.append(
                                                f"{row + 3}{self.column_number_to_letter(hygiene_col)}")
                                        break

                            # ----------------------------------------------------------
                            # 38.Kural - 60cm Vanasız ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 158 dakika olmalıdır.

                            # vanali 19.kuralda tanımlı
                            # genislik_value 37.kuralda tanımlı
                            # hygiene col 37.kuralda tanımlı
                            # 60cm genişlik için dosya isminin 4. karakterine bakılır (3 olmalı)

                            if genislik_value == 3 and "Intensive70" in current_file.file_name and not hygiene_col == -1 and not vanali:
                                # hygiene sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, hygiene_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, hygiene_col]))
                                        if not df_selected.iloc[row + 1, hygiene_col] == 158:
                                            current_file.error_elements38.append(
                                                f"{row + 3}{self.column_number_to_letter(hygiene_col)}")
                                        break

                            # ----------------------------------------------------------
                            # 39.Kural - 45cm Vanalı ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 197 dakika olmalıdır.

                            # vanali 19.kuralda tanımlı
                            # genislik_value 37.kuralda tanımlı
                            # hygiene col 37.kuralda tanımlı
                            # 45cm genişlik için dosya isminin 4. karakterine bakılır (0 olmalı)

                            if vanali and genislik_value == 0 and "Intensive70" in current_file.file_name and not hygiene_col == -1:
                                # hygiene sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, hygiene_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, hygiene_col]))
                                        if not df_selected.iloc[row + 1, hygiene_col] == 197:
                                            current_file.error_elements39.append(
                                                f"{row + 3}{self.column_number_to_letter(hygiene_col)}")
                                        break

                            # ----------------------------------------------------------
                            # 40.Kural - 45cm Vanasız ürünlerde Yoğun 70 programında Hijyen fonksiyonu seçilince kalan zaman 154 dakika olmalıdır.

                            # vanali 19.kuralda tanımlı
                            # genislik_value 37.kuralda tanımlı
                            # hygiene col 37.kuralda tanımlı
                            # 45cm genişlik için dosya isminin 4. karakterine bakılır (0 olmalı)

                            if genislik_value == 0 and not vanali and "Intensive70" in current_file.file_name and not hygiene_col == -1:
                                # hygiene sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, hygiene_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, hygiene_col]))
                                        if not df_selected.iloc[row + 1, hygiene_col] == 154:
                                            current_file.error_elements40.append(
                                                f"{row + 3}{self.column_number_to_letter(hygiene_col)}")
                                        break

                            # ----------------------------------------------------------
                            # 41.Kural - KV1000'li akışlarda TEPE pozisyonu olmamalı

                            # kv1000_check 19.kuralda tanımlı

                            if kv1000_check:
                                for row in range(2, stop_rowNumber-2):
                                    for col in range(7, len(df_selected.columns), 4):
                                        vana_values = df_selected.iloc[row, col].split("-")
                                        # vana_values[1, 3, 5, 7] kontrol edilecek
                                        if df_selected.iloc[row, col - 2] != 0:
                                            if vana_values[1] == "TEPE" or vana_values[3] == "TEPE" or vana_values[5] == "TEPE" or vana_values[7] == "TEPE":
                                                current_file.error_elements41.append(f"{deterjan_rows.index[row] + 2}{self.column_number_to_letter(col)}")

                            # ----------------------------------------------------------
                            # 42.Kural - SIRKULASYON / ISITICI , 0 harici değer , Kapalı pozisyon + S6 S7 - FAIL vermeli

                            for row in range(2, stop_rowNumber-2):
                                if df_selected.iloc[row, 2] == "SIRKULASYON / ISITICI":
                                    for col in range(4, len(df_selected.columns), 4):
                                        # 0 olmayan değerlerin kontrolü
                                        if df_selected.iloc[row, col] != 0:
                                            cvv_cells = df_selected.iloc[row, col+3].split("/")
                                            flat_list = [item for part in cvv_cells for item in part.split('-')]
                                            result = [int(item.strip()) if item.strip().isdigit() else item.strip()
                                                      for item in flat_list]
                                            # result[1, 4, 7, 10] değerlerine KAPALI kontrolü
                                            # result[2, 5, 8, 11] değerlerine S6 veya S7 kontrolü
                                            if (result[1] == "KAPALI" and result[2] == "S6") or (result[1] == "KAPALI" and result[2] == "S7"):
                                                current_file.error_elements42.append(f"{row+2}{self.column_number_to_letter(col+3)}")
                                            elif (result[4] == "KAPALI" and result[5] == "S6") or (result[4] == "KAPALI" and result[5] == "S7"):
                                                current_file.error_elements42.append(f"{row+2}{self.column_number_to_letter(col+3)}")
                                            elif (result[7] == "KAPALI" and result[8] == "S6") or (result[7] == "KAPALI" and result[8] == "S7"):
                                                current_file.error_elements42.append(f"{row+2}{self.column_number_to_letter(col+3)}")
                                            elif (result[10] == "KAPALI" and result[11] == "S6") or (result[10] == "KAPALI" and result[11] == "S7"):
                                                current_file.error_elements42.append(f"{row+2}{self.column_number_to_letter(col+3)}")
                            # ----------------------------------------------------------
                            # 43.Kural - MSC2 li ürünlerde, ilk su alma adımı 6.6 litreden büyük olmalı (mini 30 ve hızlı 58)

                            file_name = self.excel_path.split("/")[-1]  # Dosya yolundan dosya adını al

                            # Kapı açma sütununu bul

                            for col in range(0, len(df_selected.columns)):
                                if "KAPI AÇMA (KA-1)" in str(df_selected.iloc[0, col]):
                                    kapiAcma_col = col
                                    print(kapiAcma_col)
                                    break

                            # dosyanın adı ve boşaltma vanası ile MSC2 mi değil mi kontrolü

                            if "Mini30" in current_file.file_name or "Hızlı58" in current_file.file_name or "QuickDry" in current_file.file_name:
                                check_bosaltma = False
                                for row in range(sicakDurulama2_rowNumber + sicakDurulama2_adim_sayisi - 2, len(df_selected)):

                                    if "BOSALTMA VANASI" in df_selected.iloc[row, 2]:
                                        if 0 in df_selected.iloc[row, 5:: 4].values:
                                            check_bosaltma = True
                                            break
                                    else:
                                        check_bosaltma = False
                                        print("bosaltma yok")
                                        break

                                msc2_check = check_bosaltma or "MSC2" in file_name
                                if msc2_check:
                                    # "SU GİRİŞ VANASI" olan hücreleri bul
                                    row_number = df_selected.iloc[:, 2][df_selected.iloc[:, 2] == "SU GIRIS VANASI"].index.tolist()

                                    if row_number:
                                        su_giris_degeri = df_selected.iloc[row_number[0], 5]
                                        print(f"İlk 'SU GİRİŞ VANASI' değeri: {su_giris_degeri}")

                                        su_giris_degeri_kapiAcma = df_selected.iloc[row_number[0], kapiAcma_col+1]
                                        print(f"İlk kapı açma 'SU GİRİŞ VANASI' değeri: {su_giris_degeri_kapiAcma}")

                                        # Değer kontrolü
                                        if su_giris_degeri >= 6.6 or su_giris_degeri_kapiAcma >= 6.6:
                                            print("true")
                                        else:
                                            print("MSC2 su hatalı: İlk 'SU GİRİŞ VANASI' 6.6 litreden küçük.")
                                            current_file.error_elements43.append(
                                                f"HATA: 'SU GİRİŞ VANASI' değeri 6.6 litreden küçük.")
                                    else:
                                        print("HATA: 'SU GİRİŞ VANASI' bulunamadı.")
                            # ----------------------------------------------------------
                            # 44.Kural - Silence süresi max dk (normal fonksiyonun 65 dk fazlasından az olmalı)

                            file_name = self.excel_path.split("/")[-1]

                            fonksiyonsuz_col = -1
                            silence_col = -1
                            kapiAcma_col = -1
                            kapiAcma_silence_col = -1
                            tablet_col = -1
                            tablet_silence_col = -1
                            yarimyuk_col = -1
                            yarimyuk_silence_col = -1
                            zone_col = -1
                            zone_silence_col = -1
                            mft_col = -1
                            mft_silence_col = -1
                            ekstraRinse_col = -1
                            ekstraRinse_silence_col = -1
                            ekstraKurutma_col = -1
                            ekstraKurutma_silence_col = -1

                            # Fonksiyonsuz'un bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1)" in str(df_selected.iloc[0, col]):
                                    fonksiyonsuz_col = col
                                    break

                            # Silence'ın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE (S-1)" in str(df_selected.iloc[0, col]):
                                    silence_col = col
                                    break

                            # Tablet'in bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "TABLET (T-1)" in str(df_selected.iloc[0, col]):
                                    tablet_col = col
                                    break

                            # Silence + Tablet'in bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ TABLET (S-2)" in str(df_selected.iloc[0, col]):
                                    tablet_silence_col = col
                                    break

                            # 1/2'in bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "1/2 (YY-1)" in str(df_selected.iloc[0, col]):
                                    yarimyuk_col = col
                                    break

                            # Silence + 1/2'in bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ 1/2 (S-3)" in str(df_selected.iloc[0, col]):
                                    yarimyuk_silence_col = col
                                    break

                            # Zone'un bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "ZONE (Z-1)" in str(df_selected.iloc[0, col]):
                                    zone_col = col
                                    break

                            # Silence + Zone'un bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ ZONE (S-4)" in str(df_selected.iloc[0, col]):
                                    zone_silence_col = col
                                    break

                            # MFT'nin bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "MFT (MFT-1)" in str(df_selected.iloc[0, col]):
                                    mft_col = col
                                    break

                            # Silence + MFT'nin bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ MFT (S-5)" in str(df_selected.iloc[0, col]):
                                    mft_silence_col = col
                                    break

                            # Kapı Açma'nın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "KAPI AÇMA (KA-1)" in str(df_selected.iloc[0, col]):
                                    kapiAcma_col = col
                                    break

                            # Silence + kapıAçma'nın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ KAPI AÇMA (S-6)" in str(df_selected.iloc[0, col]):
                                    kapiAcma_silence_col = col
                                    break

                            # Ekstra Kurutuma'nın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "EKSTRA KURUTMA (T-1)" in str(df_selected.iloc[0, col]):
                                    ekstraKurutma_col = col
                                    break

                            # Silence + EkstraKurutma'nın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ EKSTRA KURUTMA (S-8)" in str(df_selected.iloc[0, col]):
                                    ekstraKurutma_silence_col = col
                                    break

                            # Ekstra Rinse'ın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "EKSTRA RINSE (ER-1)" in str(df_selected.iloc[0, col]):
                                    ekstraRinse_col = col
                                    break

                            # Silence + EkstraRinse'ın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "SILENCE+ EKSTRA RINSE (S-7)" in str(df_selected.iloc[0, col]):
                                    ekstraRinse_silence_col = col
                                    break

                            # Silence süresi ,fonksiyonsuz + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                if str(df_selected.iloc[row, silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(df_selected.iloc[row + 1, silence_col]))

                                    if not df_selected.iloc[row + 1, silence_col] <= df_selected.iloc[
                                        row + 1, fonksiyonsuz_col] + 68 :
                                        current_file.error_elements44.append("HATA")
                                    break

                            # Tablet/Silence süresi ,tablet + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, tablet_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, tablet_col]))
                                if str(df_selected.iloc[row, tablet_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, tablet_silence_col]))

                                    if not df_selected.iloc[row + 1, tablet_silence_col] <= df_selected.iloc[
                                        row + 1, tablet_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # 1/2/Silence süresi ,1/2 + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, yarimyuk_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, yarimyuk_col]))
                                if str(df_selected.iloc[row, yarimyuk_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, yarimyuk_silence_col]))

                                    if not df_selected.iloc[row + 1, yarimyuk_silence_col] <= df_selected.iloc[
                                        row + 1, yarimyuk_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # Zone/Silence süresi ,Zone + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, zone_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, zone_col]))
                                if str(df_selected.iloc[row, zone_silence_col]) == "Kalan Zaman":
                                    print(current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, zone_silence_col]))

                                    if not df_selected.iloc[row + 1, zone_silence_col] <= df_selected.iloc[
                                        row + 1, zone_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # MFT/Silence süresi ,MFT + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, mft_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, mft_col]))
                                if str(df_selected.iloc[row, mft_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, mft_silence_col]))

                                    if not df_selected.iloc[row + 1, mft_silence_col] <= df_selected.iloc[
                                        row + 1, mft_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # Kapı Açma/Silence süresi ,Kapı açma + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, kapiAcma_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, kapiAcma_col]))
                                if str(df_selected.iloc[row, kapiAcma_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, kapiAcma_silence_col]))

                                    if not df_selected.iloc[row + 1, kapiAcma_silence_col] <= df_selected.iloc[
                                        row + 1, kapiAcma_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # EkstraKurutma/Silence süresi ,EkstraKurutma + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, ekstraKurutma_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, ekstraKurutma_col]))
                                if str(df_selected.iloc[row, ekstraKurutma_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, ekstraKurutma_silence_col]))

                                    if not df_selected.iloc[row + 1, ekstraKurutma_silence_col] <= df_selected.iloc[
                                        row + 1, ekstraKurutma_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # EkstraRinse/Silence süresi ,EkstraRinse + 68'den az olmalı

                            for row in range(0, len(df_selected)):
                                if str(df_selected.iloc[row, ekstraRinse_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, ekstraRinse_col]))
                                if str(df_selected.iloc[row, ekstraRinse_silence_col]) == "Kalan Zaman":
                                    print(
                                        current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, ekstraRinse_silence_col]))

                                    if not df_selected.iloc[row + 1, ekstraRinse_silence_col] <= \
                                           df_selected.iloc[row + 1, ekstraRinse_col] + 68:
                                        current_file.error_elements44.append("HATA")
                                    break

                            # ----------------------------------------------------------
                            # 45.Kural - fonksiyonsuz ile fonk + kapı açma süresi aynı olmalı

                            file_name = self.excel_path.split("/")[-1]

                            if "Mini30" in current_file.file_name or "QuickDry" in current_file.file_name:

                                fonksiyonsuz_col = -1
                                kapiAcma_col = -1

                                # Fonksiyonsuz'un bulunduğu sütunu bul
                                for col in range(0, len(df_selected.columns)):
                                    if "FONKSİYONSUZ (N-1)" in str(df_selected.iloc[0, col]):
                                        fonksiyonsuz_col = col
                                        break

                                # Kapı Açma'nın bulunduğu sütunu bul
                                for col in range(0, len(df_selected.columns)):
                                    if "KAPI AÇMA (KA-1)" in str(df_selected.iloc[0, col]):
                                        kapiAcma_col = col
                                        break

                                # Fonksiyonsuz ve kapı açma süreleri eşit olmalı

                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, kapiAcma_col]))

                                        if not df_selected.iloc[row + 1, fonksiyonsuz_col] == df_selected.iloc[row + 1, kapiAcma_col]:
                                            current_file.error_elements45.append("HATA")
                                        break
                            # ----------------------------------------------------------
                            # 46.Kural - makinetemizleme süreleri doğru girilmelidir

                            file_name = self.excel_path.split("/")[-1]
                            fonksiyonsuz_col = -1
                            kapiAcma_col = -1

                            # Fonksiyonsuz'un bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1)" in str(df_selected.iloc[0, col]):
                                    fonksiyonsuz_col = col
                                    break

                            # Kapı Açma'nın bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "KAPI AÇMA (KA-1)" in str(df_selected.iloc[0, col]):
                                    kapiAcma_col = col
                                    break

                            # Makine Temizleme Kalan Zaman 15 dk
                            if "MakineTemizleme" in current_file.file_name and not fonksiyonsuz_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                        if not df_selected.iloc[row + 1, fonksiyonsuz_col] == 75:
                                            current_file.error_elements46.append(
                                                f"{row + 3}{self.column_number_to_letter(fonksiyonsuz_col)}")
                                        break

                            # Makine Temizleme Kalan Zaman 15 dk - Kapı Açma
                            if "MakineTemizleme" in current_file.file_name and not kapiAcma_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, kapiAcma_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, kapiAcma_col]))
                                        if not df_selected.iloc[row + 1, kapiAcma_col] == 75:
                                            current_file.error_elements46.append(
                                                f"{row + 3}{self.column_number_to_letter(kapiAcma_col)}")
                                        break
                            # ----------------------------------------------------------
                            # 47.Kural - ön yıkama süreleri doğru girilmelidir
                            # Ön yıkama Kalan Zaman 10 dk
                            if "PreWash" in current_file.file_name and not fonksiyonsuz_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                        if not df_selected.iloc[row + 1, fonksiyonsuz_col] == 15:
                                            current_file.error_elements47.append(
                                                f"{row + 3}{self.column_number_to_letter(fonksiyonsuz_col)}")
                                        break

                            # Ön yıkama Kalan Zaman 10 dk - Kapı Açma
                            if "PreWash" in current_file.file_name and not kapiAcma_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, kapiAcma_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, kapiAcma_col]))
                                        if not df_selected.iloc[row + 1, kapiAcma_col] == 15:
                                            current_file.error_elements47.append(
                                                f"{row + 3}{self.column_number_to_letter(kapiAcma_col)}")
                                        break
                            # ----------------------------------------------------------
                            # 48.Kural - mini30 süreleri doğru girilmelidir
                            # Mini 30 Kalan Zaman 30 dk - Fonksiyonsuz
                            if "Mini30" in current_file.file_name and not fonksiyonsuz_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                        if not (df_selected.iloc[row + 1, fonksiyonsuz_col] == 30):
                                            current_file.error_elements48.append(
                                                f"{row + 3}{self.column_number_to_letter(fonksiyonsuz_col)}")
                                        break


                            # Mini 30 Kalan Zaman 30 dk - Kapı Açma
                            if "Mini30" in current_file.file_name and not kapiAcma_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, kapiAcma_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, kapiAcma_col]))
                                        if not df_selected.iloc[row + 1, kapiAcma_col] == 30:
                                            current_file.error_elements48.append(
                                                f"{row + 3}{self.column_number_to_letter(kapiAcma_col)}")
                                        break
                            # ----------------------------------------------------------
                            # 49.Kural - hızlı58 süreleri doğru girilmelidir
                            # Hızlı58 Kalan Zaman 58 dk
                            if "Hızlı58" in current_file.file_name and not fonksiyonsuz_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, fonksiyonsuz_col]))
                                        if not df_selected.iloc[row + 1, fonksiyonsuz_col] == 58:
                                            current_file.error_elements49.append(
                                                f"{row + 3}{self.column_number_to_letter(fonksiyonsuz_col)}")
                                        break

                            # Hızlı58 Kalan Zaman 58 dk - Kapı Açma
                            if "Hızlı58" in current_file.file_name and not kapiAcma_col == -1:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, kapiAcma_col]) == "Kalan Zaman":
                                        print(
                                            current_file.file_name + " " + str(
                                                df_selected.iloc[row + 1, kapiAcma_col]))
                                        if not df_selected.iloc[row + 1, kapiAcma_col] == 58:
                                            current_file.error_elements49.append(
                                                f"{row + 3}{self.column_number_to_letter(kapiAcma_col)}")
                                        break

                                        # ----------------------------------------------------------
                            # 50.Kural - Auto programlarda 0 kir,yarım kir,tam kir ve 3/2 kir süreleri eşit olmalı

                            fonksiyonsuz_sifirYuk_col = -1
                            fonksiyonsuz_yarimYuk_col = -1
                            fonksiyonsuz_tamYuk_col = -1
                            fonksiyonsuz_yogunYuk_col = -1

                            # Fonksiyonsuz 0 yük'ün bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1) ( 0 Kir )" in str(df_selected.iloc[0, col]) :
                                    fonksiyonsuz_sifirYuk_col = col
                                    break

                            # Fonksiyonsuz yarım yük'ün bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1) ( Yarım Kir )" in str(df_selected.iloc[0, col]):
                                    fonksiyonsuz_yarimYuk_col = col
                                    break

                            # Fonksiyonsuz tam yük'ün bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1) ( Tam Kir )" in str(df_selected.iloc[0, col]):
                                    fonksiyonsuz_tamYuk_col = col
                                    break

                            # Fonksiyonsuz yogun yük'ün bulunduğu sütunu bul
                            for col in range(0, len(df_selected.columns)):
                                if "FONKSİYONSUZ (N-1) ( Yoğun Kir )" in str(df_selected.iloc[0, col]):
                                    fonksiyonsuz_yogunYuk_col = col
                                    break

                            # dosya ismi AUTO içeriyor mu kontrolü
                            auto = False
                            if "Auto" in current_file.file_name :
                                auto = True

                            if auto:
                                # Fonksiyonsuz sütunu kalan zamanı kontrol etme
                                for row in range(0, len(df_selected)):
                                    if str(df_selected.iloc[row, fonksiyonsuz_sifirYuk_col]) == "Kalan Zaman":
                                        print(current_file.file_name + " " + str(df_selected.iloc[row + 1, fonksiyonsuz_sifirYuk_col]))

                                    if str(df_selected.iloc[row, fonksiyonsuz_yarimYuk_col]) == "Kalan Zaman":
                                        print(current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, fonksiyonsuz_yarimYuk_col]))

                                    if str(df_selected.iloc[row, fonksiyonsuz_tamYuk_col]) == "Kalan Zaman":
                                        print(current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, fonksiyonsuz_tamYuk_col]))

                                    if str(df_selected.iloc[row, fonksiyonsuz_yogunYuk_col]) == "Kalan Zaman":
                                        print(current_file.file_name + " " + str(
                                            df_selected.iloc[row + 1, fonksiyonsuz_yogunYuk_col]))

                                        if not df_selected.iloc[row + 1, fonksiyonsuz_sifirYuk_col] == df_selected.iloc[row + 1, fonksiyonsuz_yarimYuk_col] == df_selected.iloc[row + 1, fonksiyonsuz_tamYuk_col] == df_selected.iloc[row + 1, fonksiyonsuz_yogunYuk_col] :
                                            current_file.error_elements50.append("HATA")
                                        break
                            else:
                                print("auto yok")

                            self.compared_excel = True
                            self.compared_msg.config(text="Kontrol Tamamlandı. Hata Dosyası oluşturmak için butona tıklayın.")

                    except Exception as e:
                        print("Hata:", e)
                        traceback.print_exc()
        else:
            print("Önce bir Excel dosyası seçmelisiniz.")

    def create_excel(self):
        if self.compared_excel:

            all_data = []

            for file in self.exceles:
                for i in range(1, 50): # Kural sayısı değişirse düzelt
                    error_elements_attr = f"error_elements{i}"
                    kural_attr = f"kural{i}"

                    if hasattr(file, error_elements_attr) and hasattr(self, kural_attr):
                        error_elements = getattr(file, error_elements_attr)
                        kural = getattr(self, kural_attr)
                        if error_elements:
                            hatali_hucreler = ', '.join(map(str, error_elements))
                            all_data.append([file.file_name, kural, hatali_hucreler])
                all_data.append(["", "", ""])

            if all_data:
                df = pd.DataFrame(all_data, columns=['Dosya Adı', 'Kural', 'Hatalı Hücreler'])
                directory_path = filedialog.askdirectory()
                file_name_raw = "akış_fail_cells"
                file_name = "akış_fail_cells.xlsx"
                full_path = os.path.join(directory_path, file_name)

                # Check if file exists and generate a unique name if necessary
                counter = 1
                while os.path.exists(full_path):
                    full_path = os.path.join(directory_path, f"{file_name_raw}_{counter}.xlsx")
                    counter += 1

                df.to_excel(full_path, index=False)

                # hata excelinin sütun genişliklerini ayarlama
                wb = load_workbook(full_path)
                ws = wb.active

                column_widths = [50, 70, 50]
                for i, column_width in enumerate(column_widths, start=1):
                    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = column_width

                wb.save(full_path)

                print(f"Hata hücreleri başarılı bir şekilde yazıldı: {full_path}")
                os.system(f'start "" "{full_path}"')

        else:
            print("Önce bir Excel dosyası karşılaştırmalısınız.")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = InterFace()
    app.run()
